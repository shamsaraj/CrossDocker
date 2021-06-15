print"################################################################################"
print "CrossDocker                                                                   "
print "A tool for performing easy cross-docking using AutoDock Vina and data collection"
print "v. 1.0                                                                   "
print "Copyright (C) 2016  Jamal Shamsara                                       "
print "e-mail: shamsaraj@mums.ac.ir; jshamsara@yahoo.com"
print "################################################################################"
print "Usage of CrossDocker is free without any limitations."
print "  There is NO warranty"
print "################################################################################"
print "Please see the manual and config files for more information"
print "http://www.pharm-sbg.com"
print"#################################################################################"
print "..."
import os
import fileinput
import shutil
import errno
import csv
from rmsd import rmsd2
from center import center_of_mass as center
from openpyxl import Workbook
import numpy
from openpyxl.styles import Color, PatternFill, Font, Border, Style
from openpyxl.formatting import ColorScaleRule, CellIsRule, FormulaRule
from ConfigParser import SafeConfigParser
import sys
from prepare_ligand4_m2 import PL
from prepare_receptor4_m2 import PR
from time import sleep
sleep (5)
##########################################################Functions########################################################
def is_number(s):
    try:
        float(s)
        return True
    except ValueError:
        return False
def makefile(name, path, text):
    os.chdir(path)
    newfile = open(name, "w")
    newfile.write(text)
    newfile.close()
def make_directory_if_not_exists(path):
    while not os.path.isdir(path):
        try:
            os.makedirs(path)
            break
        except OSError as exception:
            if exception.errno != errno.EEXIST:
                raise
        except WindowsError:
            print "got WindowsError"
            pass
def chk_dir(path,name):
    if not os.path.isdir(path):
        print "#######################################################################################"
        print " The path for the " +  name + " is not correctly defined in config.txt file"
        print "#######################################################################################"
def copy_files(src, dest, ext):
    src_files = os.listdir(src)
    for file_name in src_files:
        if file_name.endswith(ext):
            full_file_name = os.path.join(src, file_name)
            if (os.path.isfile(full_file_name)):
                shutil.copy2(full_file_name, dest)
def extraction(keyword, datapath, output):
    os.chdir(datapath)
    ff = open(output, "w")
    out = csv.writer(ff)
    out.writerow(["name", "energy"])
    for filename in os.listdir(datapath):
        if filename.endswith(".pdbqt"):
            f = open(filename, "r")
            # filename2=filename[7:-6]
            filename2 = filename[:-6]
            for line in f:
                if keyword in line:
                    words = line.split()
                    for word in words:
                        answer = is_number(word)
                        if answer == True:
                            if abs(float(word)) > 1:
                                out.writerow([filename2, word])
def extraction2 (datapath, name2, name, keyword, number):
    os.chdir(datapath)
    for filename in os.listdir(datapath):
        if filename.endswith(name):
            if filename.startswith(name2):
                f = open(filename, "r")
                for line in f:
                    if keyword in line:
                        words1 = line.split()
                        temp = f.next()
                        words2 = temp.split()
                        if words1[1]== str(number):
                            return str(words2[3])
def we_are_frozen():
    """Returns whether we are frozen via py2exe.
    This will affect how we find out where we are located."""
    return hasattr(sys, "frozen")
def module_path():
    """ This will get us the program's directory,
    even if we are frozen using py2exe"""
    if we_are_frozen():
        return os.path.dirname(unicode(sys.executable, sys.getfilesystemencoding( )))
    return os.path.dirname(unicode(__file__, sys.getfilesystemencoding( )))
def receptor2pdbqt (RF , OF):
	PR (RF, OF)
def ligand2pdbqt (LF):
	PL (LF)
##########################################################################################################################
config = SafeConfigParser()
config.read('config.txt')
CDIR= module_path()
N = 500 #max number of receptors
NN = 20 #max number of poses per ligand
recfile = [""]*N
newpath = [""]*N
wb = Workbook()
wb2 = Workbook()
wb3 = Workbook()
wb4 = Workbook()
wb5 = Workbook()
wb6 = Workbook()
BoldFont= Font(bold = True)
styleObj = Style(font=BoldFont)
RMSD = numpy.zeros((N,N,NN))
Energy = numpy.zeros((N,N,NN))
for h in range (0,NN):
    for m in range (0,N):
        for n in range (0,N):
            RMSD [m,n,h]= 50 # This should be higher than any obtained RMSD
Filename = numpy.empty((N,N),dtype="S4")
i=0
#Babel_path = config.get("paths", 'RpathE') # '"D:/Program Files/R/R-3.0.2/bin/x64/Rscript.exe"' # Path for Rscript.exe
# For Windows:
orig_path = config.get("paths", 'project_path') + "/" #"D:/test3" # output path - will be crated by the script
recpath = config.get("paths", 'user_receptors_path') + "/" # "D:/project/"  # input path
ligpath = config.get("paths", 'user_ligands_path') + "/"
script_path = CDIR # path for vina.exe autodockscripts
path = orig_path
chk_dir(path, "project_path")
chk_dir(recpath, "user_receptors_path")
chk_dir(ligpath, "user_ligands_path")
os.chdir(path)
# Receptors preparation
for filename_r in os.listdir(recpath):
    if filename_r.endswith("pdb")or filename_r.endswith("mol2"):
        recfile[i] = recpath + filename_r
        newpath[i] = path + filename_r[:-4] + "/"
        make_directory_if_not_exists(newpath[i])
        os.chdir(recpath)
        if filename_r.endswith("pdb"):
            out = recfile[i] + "qt"
        elif filename_r.endswith("mol2"):
            out = recfile[i][:-5] + "pdbqt"
        receptor2pdbqt (recfile[i] , out)
         # ZN CA charges
        for line in fileinput.input(recfile[i] + "qt", inplace=1):
            line = line.replace("0.000 Zn", "2.000 Zn")
            line = line.replace("0.000 Ca", "2.000 Ca")
            line = line.replace("0.000 Mg", "2.000 Mg")
            print (line)
        # Makin the conf file
        gauss1 = -0.035579  # -0.035579
        gauss2 = -0.005156  # -0.005156
        repulsion = 0.84024500000000002  # 0.84024500000000002
        hydrophobic = -0.035069000000000003  # -0.035069000000000003
        hydrogen = -0.58743900000000004  # -0.58743900000000004
        rot = 0.058459999999999998  # 0.058459999999999998
        for filename1 in os.listdir(ligpath):
            if filename1.endswith("pdb")and filename1[0:4]== filename_r[0:4]:
                get_position = center (ligpath + filename1)  ###########################
        center_x = str(get_position[0])
        center_y = str(get_position[1])
        center_z = str(get_position[2])
        size_x = str (config.get("values", 'size_x'))
        size_y = str (config.get("values", 'size_y'))
        size_z = str (config.get("values", 'size_z'))
        exhaustiveness = str (config.get("values", 'exhaustiveness'))
        num_modes = str (config.get("values", 'num_modes'))
        energy_range = str(3)
        makefile("conf.txt", newpath[i],
                 "receptor = " + recfile[i] + "qt" + "\nnum_modes = " + num_modes + "\nenergy_range = " + energy_range + "\nexhaustiveness = " + exhaustiveness + "\ncenter_x =  " + center_x + "\ncenter_y =  " + center_y + "\ncenter_z =  " + center_z + "\nsize_x =  " + size_x + "\nsize_y =  " + size_y + "\nsize_z =  " + size_z + "\nweight_gauss1 = " + str(
                     gauss1) + " \nweight_gauss2 = " + str(gauss2) + " \nweight_repulsion = " + str(
                     repulsion) + " \nweight_hydrophobic = " + str(hydrophobic) + " \nweight_hydrogen = " + str(
                     hydrogen) + " \nweight_rot = " + str(rot))
        if i == 0:
            # Ligands preparation
            for filename in os.listdir(ligpath):
                if filename.endswith("pdb") or filename.endswith("mol2"):
                    ligfile = ligpath + filename
                    os.chdir(ligpath)
                    ligand2pdbqt(ligfile)
            # randomize input
            for filename in os.listdir(ligpath):
                if filename.endswith("pdbqt")and not filename.endswith("_random.pdbqt"):
                    os.chdir(CDIR)
                    ligfile = ligpath + filename
                    cmd= "vina --ligand " + ligfile + " --config " + newpath[i] + "conf.txt --randomize_only --out " + ligfile [:-6] + "_random.pdbqt"
                    os.system(cmd)
            # make xyz files from inputs
            for filename in os.listdir(ligpath):
                 if filename.endswith("pdbqt")and not filename.endswith("_random.pdbqt"):###################################################
                    ligfile = ligpath + filename
                    cmd9 = "babel -ipdbqt " + ligfile + " -oxyz " + ligfile + ".xyz -h -m"
                    os.system(cmd9)
        #Dock
        for filename in os.listdir(ligpath):
            if filename.endswith("_random.pdbqt"):
                os.chdir(CDIR)
                ligfile = ligpath + filename
                print "*******************************"
                print "Ligand -------> " + ligfile
                print "Receptor -----> " + recfile[i] + "qt"
                cmd2 = "vina --ligand " + ligfile + " --config " + newpath[i] + "conf.txt " + " --out " + newpath[i] +  filename [:-6] + "_out.pdbqt"
                os.system(cmd2)
        #make xyz from output - pdbqt to xyz
        for filename in os.listdir(newpath[i]):
            if filename.endswith("_out.pdbqt"):
                ligfile = newpath[i] + filename
                cmd8 = "babel -ipdbqt " + ligfile + " -oxyz " + ligfile + ".xyz -h -m"
                os.system(cmd8)
        if i == 0:
            #output files
            ws = wb.active
            ws2 = wb2.active
            ws3 = wb3.active
            ws4 = wb4.active
            ws5 = wb5.active
            ws6 = wb6.active
            ws.append (["receptor" , "ligand", "RMSD", "Energy"])
            ws2.append (["receptor" , "ligand", "Best RMSD", "Energy"])
            ws3["A1"] = ""
            ws4["A1"] = ""
            ws5["A1"] = ""
            ws6["A1"] = ""
        #RMSD calculation
        j = -1
        for filename1 in os.listdir(ligpath):
            if filename1.endswith("-l.pdbqt1.xyz"):
                j = j + 1
                for filename2 in os.listdir(newpath[i]):
                    if filename2.endswith(".xyz"):
                        if filename1[0:4]== filename2[0:4]:####
                            k = filename2[-5:-4]
                            if filename2[-6:-5]== "1" or filename2[-6:-5] == "2":
                                k = filename2[-6:-4]
                            energy = extraction2 (newpath[i], filename1[0:4], "_random_out.pdbqt", "MODEL", k)
                            RMSD [i,j,int(k)-1] = rmsd2(ligpath + filename1, newpath[i]+ filename2)
                            Energy [i,j,int(k)-1] = energy
                            Filename[i,j] = filename2 [:4]
                            ws.append([newpath[i], filename2, rmsd2(ligpath + filename1, newpath[i]+ filename2), float(energy)])
                wb.save(orig_path+"/Output.xlsx")
                low = min (RMSD[i,j])
                Energy_best = Energy[i,j,0]
                location = numpy.where(RMSD[i,j]==low)
                Energy_cor = Energy[i,j,location[0][0]]
                rmsd_energy1 =  RMSD [i,j,0]
                #ws2 = wb2.active
                ws2.append([newpath[i], Filename[i,j], low, Energy_cor])
                wb2.save(orig_path+"/Output_the_best_RMSD.xlsx")
                ws3.cell(column=i+2, row=1, value=filename_r[:-4])
                ws3.cell(column=1, row=j+2, value=Filename[i,j])
                ws3.cell(column=i+2, row=j+2, value=low)
                wb3.save(orig_path+"/Table_the_best_RMSD.xlsx")
                ws4.cell(column=i+2, row=1, value=filename_r[:-4])
                ws4.cell(column=1, row=j+2, value=Filename[i,j])
                ws4.cell(column=i+2, row=j+2, value=Energy_cor)
                wb4.save(orig_path+"/Table_energy_for_the_best_RMSD.xlsx")
                ws5.cell(column=i+2, row=1, value=filename_r[:-4])
                ws5.cell(column=1, row=j+2, value=Filename[i,j])
                ws5.cell(column=i+2, row=j+2, value=Energy_best)
                wb5.save(orig_path+"/Table_the_best_energy.xlsx")
                ws6.cell(column=i+2, row=1, value=filename_r[:-4])
                ws6.cell(column=1, row=j+2, value=Filename[i,j])
                ws6.cell(column=i+2, row=j+2, value=rmsd_energy1)
                wb6.save(orig_path+"/Table_RMSD_for_the_best_energy.xlsx")
        i = i+ 1
wb.save(orig_path+"/Output.xlsx")
for jj in range(2,i+2):
    ws3.cell(column=jj, row=jj).style=styleObj
    c1= ws3.cell(column=jj, row=2)
    c1 = c1.coordinate
    c2= ws3.cell(column=jj, row=i+1)
    c2 = c2.coordinate
    if jj == 2:
        C1=c1
    ws3.cell(column=1, row=i+2, value= "Average")
    ws3.cell(column=1, row=i+3, value= "RMSD < 2.0")
    ws3.cell(column=jj, row=i+2, value= "=AVERAGE(" + c1 + ":" + c2 + ")")
    mylist = [0]*i
    for ii in range (2, i+2):
        mylist[ii-2]= ws3.cell(column=jj, row=ii).value
    count= sum(1 for x in mylist if float(x) <= 2.0)
    ws3.cell(column=jj, row=i+3, value= count)
ws3.conditional_formatting.add(C1 + ":" + c2,
               ColorScaleRule(start_type='percentile', start_value=10, start_color=Color('44f628'),
                           mid_type='percentile', mid_value=50, mid_color=Color('f0f228'),
                           end_type='percentile', end_value=90, end_color=Color('f66767')))
wb3.save(orig_path+"/Table_the_best_RMSD.xlsx")
for jj in range(2,i+2):
    c1= ws4.cell(column=jj, row=2)
    c1 = c1.coordinate
    c2= ws4.cell(column=jj, row=i+1)
    c2 = c2.coordinate
    if jj == 2:
        C1=c1
    ws4.cell(column=1, row=i+2, value= "Average")
    ws4.cell(column=jj, row=i+2, value= "=AVERAGE(" + c1 + ":" + c2 + ")")
ws4.conditional_formatting.add(C1 + ":" + c2,
               ColorScaleRule(start_type='percentile', start_value=10, start_color=Color('44f628'),
                           mid_type='percentile', mid_value=50, mid_color=Color('f0f228'),
                           end_type='percentile', end_value=90, end_color=Color('f66767')))
wb4.save(orig_path+"/Table_energy_for_the_best_RMSD.xlsx")
for jj in range(2,i+2):
    c1= ws5.cell(column=jj, row=2)
    c1 = c1.coordinate
    c2= ws4.cell(column=jj, row=i+1)
    c2 = c2.coordinate
    if jj == 2:
        C1=c1
    ws5.cell(column=1, row=i+2, value= "Average")
    ws5.cell(column=jj, row=i+2, value= "=AVERAGE(" + c1 + ":" + c2 + ")")
ws5.conditional_formatting.add(C1 + ":" + c2,
               ColorScaleRule(start_type='percentile', start_value=10, start_color=Color('44f628'),
                           mid_type='percentile', mid_value=50, mid_color=Color('f0f228'),
                           end_type='percentile', end_value=90, end_color=Color('f66767')))
wb5.save(orig_path+"/Table_the_best_energy.xlsx")
for jj in range(2,i+2):
    c1= ws6.cell(column=jj, row=2)
    c1 = c1.coordinate
    c2= ws6.cell(column=jj, row=i+1)
    c2 = c2.coordinate
    if jj == 2:
        C1=c1
    ws6.cell(column=1, row=i+2, value= "Average")
    ws6.cell(column=1, row=i+3, value= "RMSD < 2.0")
    ws6.cell(column=jj, row=i+2, value= "=AVERAGE(" + c1 + ":" + c2 + ")")
    mylist = [0]*i
    for ii in range (2, i+2):
        mylist[ii-2]= ws6.cell(column=jj, row=ii).value
    count= sum(1 for x in mylist if float(x) <= 2.0)
    ws6.cell(column=jj, row=i+3, value= count)
ws6.conditional_formatting.add(C1 + ":" + c2,
               ColorScaleRule(start_type='percentile', start_value=10, start_color=Color('44f628'),
                           mid_type='percentile', mid_value=50, mid_color=Color('f0f228'),
                           end_type='percentile', end_value=90, end_color=Color('ff1d1d')))
wb6.save(orig_path+"/Table_RMSD_for_the_best_energy.xlsx")
#f66767
